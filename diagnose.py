"""
BYOD Database Diagnostic Tool
This script checks your Excel file and shows exactly what the automation script sees
i used this to diagnose why the automation script was showing 0 registrations to process,
even though there was data in the Excel file. It helps identify issues like: 
1. File not found
2. Wrong sheet name
3. Data starting from wrong row
4. Status column not set to 'Pending' or empty
5. Excel file not saved after adding data. 
Might be useful for you if you are facing similar issues. Just run: python diagnose.py
"""

import openpyxl
import os

def diagnose_excel_file():
    excel_file = 'NITDA_BYOD_Database.xlsx'
    
    print("\n" + "="*70)
    print("NITDA BYOD DATABASE DIAGNOSTIC")
    print("="*70)
    
    # Check if file exists
    print("\n1. FILE CHECK")
    print("-" * 70)
    if not os.path.exists(excel_file):
        print(f"❌ ERROR: {excel_file} not found in current directory!")
        print(f"Current directory: {os.getcwd()}")
        print("\nFiles in current directory:")
        for f in os.listdir('.'):
            print(f"  - {f}")
        return
    else:
        print(f"✓ Found: {excel_file}")
        file_size = os.path.getsize(excel_file)
        print(f"✓ File size: {file_size:,} bytes")
    
    # Open workbook
    print("\n2. WORKBOOK CHECK")
    print("-" * 70)
    try:
        wb = openpyxl.load_workbook(excel_file)
        print(f"✓ Excel file opened successfully")
        print(f"✓ Sheet names: {wb.sheetnames}")
    except Exception as e:
        print(f"❌ ERROR opening Excel file: {e}")
        return
    
    # Check Device Registrations sheet
    print("\n3. DEVICE REGISTRATIONS SHEET")
    print("-" * 70)
    if 'Device Registrations' not in wb.sheetnames:
        print("❌ ERROR: 'Device Registrations' sheet not found!")
        print(f"Available sheets: {wb.sheetnames}")
        return
    
    sheet = wb['Device Registrations']
    print(f"✓ Sheet found: 'Device Registrations'")
    print(f"✓ Max row: {sheet.max_row}")
    print(f"✓ Max column: {sheet.max_column}")
    
    # Check headers
    print("\n4. COLUMN HEADERS (Row 1)")
    print("-" * 70)
    headers = []
    for col in range(1, min(sheet.max_column + 1, 20)):
        header_value = sheet.cell(row=1, column=col).value
        headers.append(header_value)
        print(f"Column {col:2d} ({chr(64+col)}): {header_value}")
    
    # Check for data rows
    print("\n5. DATA ROWS")
    print("-" * 70)
    data_count = 0
    
    for row in range(2, sheet.max_row + 1):
        reg_id = sheet.cell(row=row, column=1).value  # Column A
        name = sheet.cell(row=row, column=3).value    # Column C
        status = sheet.cell(row=row, column=15).value # Column O
        
        if reg_id or name:  # If there's any data
            data_count += 1
            print(f"\nRow {row}:")
            print(f"  Registration ID (A): '{reg_id}'")
            print(f"  Name (C): '{name}'")
            print(f"  Status (O): '{status}'")
            
            # Show all columns with data
            print(f"  All columns:")
            for col in range(1, min(sheet.max_column + 1, 20)):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value:
                    print(f"    {chr(64+col)}{row}: '{cell_value}'")
    
    if data_count == 0:
        print("❌ NO DATA FOUND in rows 2 and below!")
        print("\nPossible issues:")
        print("  1. Data is in wrong sheet")
        print("  2. Data starts from wrong row")
        print("  3. Excel file wasn't saved after adding data")
    else:
        print(f"\n✓ Found {data_count} rows with data")
    
    # Check what automation script would see
    print("\n6. AUTOMATION SCRIPT LOGIC TEST")
    print("-" * 70)
    
    pending_count = 0
    approved_count = 0
    
    for row in sheet.iter_rows(min_row=2, values_only=False):
        reg_id = row[0].value  # Column A (index 0)
        
        if not reg_id:
            continue
        
        status = row[14].value if len(row) > 14 else None  # Column O (index 14)
        
        print(f"\nRow {row[0].row}:")
        print(f"  Reg ID: '{reg_id}'")
        print(f"  Status (column O, index 14): '{status}'")
        
        # This is the logic from byod_automation.py
        if status == 'Pending' or not status:
            pending_count += 1
            print(f"  → Would be PROCESSED (status is Pending or empty)")
        elif status == 'Approved':
            approved_count += 1
            print(f"  → Would be sent to IT inspection")
        else:
            print(f"  → Would be SKIPPED (status: {status})")
    
    print("\n" + "-" * 70)
    print(f"SUMMARY:")
    print(f"  Rows that would be processed as NEW: {pending_count}")
    print(f"  Rows that would be sent to IT: {approved_count}")
    print(f"  Total data rows: {data_count}")
    
    # Final diagnosis
    print("\n7. DIAGNOSIS")
    print("="*70)
    
    if data_count == 0:
        print("❌ PROBLEM: No data found in Excel file")
        print("\nSOLUTION:")
        print("  1. Make sure you added data to 'Device Registrations' sheet")
        print("  2. Make sure you SAVED the Excel file (Ctrl+S)")
        print("  3. Close Excel if it's open (file might be locked)")
        
    elif pending_count == 0:
        print("❌ PROBLEM: Data exists but nothing to process")
        print("\nPossible reasons:")
        print("  1. Status column (O) is not 'Pending' or empty")
        print("  2. Registration ID column (A) is empty")
        
        print("\nSOLUTION:")
        print("  - Set Status column (O) to 'Pending' or leave it empty")
        print("  - Make sure Registration ID column (A) has values")
        
    else:
        print(f"✓ LOOKS GOOD! Should process {pending_count} registrations")
        print("\nIf automation still shows 0:")
        print("  1. Make sure Excel file is closed")
        print("  2. Make sure you saved the file")
        print("  3. Try running: python byod_automation.py")
    
    print("\n" + "="*70)


def check_current_directory():
    """Check what files are in the current directory"""
    print("\nFILES IN CURRENT DIRECTORY:")
    print("-" * 70)
    
    required_files = [
        'NITDA_BYOD_Database.xlsx',
        'byod_automation.py',
        'sync_sheets.py'
    ]
    
    current_files = os.listdir('.')
    
    for req_file in required_files:
        if req_file in current_files:
            print(f"✓ {req_file}")
        else:
            print(f"❌ {req_file} - NOT FOUND!")
    
    print("\nAll files in directory:")
    for f in sorted(current_files):
        if not f.startswith('.'):
            print(f"  - {f}")


if __name__ == "__main__":
    check_current_directory()
    diagnose_excel_file()
    
    print("\n\nNEXT STEPS:")
    print("-" * 70)
    print("1. Review the output above")
    print("2. Fix any issues identified")
    print("3. Run: python byod_automation.py")
    print()
