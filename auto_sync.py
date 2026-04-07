"""
Google Sheets Auto-Sync Script
This script automatically syncs Google Forms responses to Excel
Runs continuously in the background
"""
import sys
import gspread
from google.oauth2.service_account import Credentials
import openpyxl
from datetime import datetime
import time
import os

######added automation for excel changes start
class CombinedWatcher:
    def __init__(self, sync_obj, excel_file):
        self.sync_obj = sync_obj
        self.excel_file = excel_file
        self.last_modified = 0
         # Initialize with current modified time
        try:
            self.last_modified = os.path.getmtime(self.excel_file)
        except:
            pass
        
    def check_excel_changes(self):
        """Check if Excel was manually modified (returns True if modified)"""
        try:
            current_modified = os.path.getmtime(self.excel_file)
            
            if current_modified > self.last_modified:
                self.last_modified = current_modified
                return True
            return False
        except:
            return False
####added automation for excel changes end

class GoogleSheetsAutoSync:
    def __init__(self, credentials_file, spreadsheet_name, excel_file):
        """
        Initialize the auto-sync system
        
        Args:
            credentials_file: Path to service account JSON file
            spreadsheet_name: Name of your Google Sheet
            excel_file: Path to Excel database
        """
        self.excel_file = excel_file
        self.last_sync_time = None
        
        # Set up Google Sheets API
        SCOPES = [
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/drive'
        ]
        
        self.creds = Credentials.from_service_account_file(
            credentials_file, scopes=SCOPES)
        
        self.client = gspread.authorize(self.creds)
        
        # Open the Google Sheet
        try:
            self.sheet = self.client.open(spreadsheet_name).sheet1
            print(f"✓ Connected to Google Sheet: {spreadsheet_name}")
        except Exception as e:
            print(f"❌ Error connecting to Google Sheet: {e}")
            print("\nMake sure:")
            print("1. The spreadsheet name is correct")
            print("2. You've shared the sheet with the service account email")
            raise
    
    def get_last_row_from_excel(self):
        """Get the number of rows already in Excel"""
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            sheet = wb['Device Registrations']
            
            # Count rows with data (excluding header)
            count = 0
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row=row, column=1).value:  # If Registration ID exists
                    count += 1
            
            wb.close()
            return count
        except Exception as e:
            print(f"Error reading Excel: {e}")
            return 0
    
    def sync_new_responses(self):
        """Sync only new responses from Google Sheets to Excel"""
        try:
            # Get all Google Form responses
            all_values = self.sheet.get_all_values()
            
            if len(all_values) <= 1:  # Only headers, no data
                print("No responses in Google Sheet yet")
                return 0
            
            headers = all_values[0]
            data_rows = all_values[1:]  # Skip header row
            
            # Get current Excel row count
            excel_row_count = self.get_last_row_from_excel()
            
            # Calculate new rows
            new_rows = data_rows[excel_row_count:]
            
            if not new_rows:
                return 0  # No new data
            
            # Load Excel
            wb = openpyxl.load_workbook(self.excel_file)
            sheet = wb['Device Registrations']
            
            # Get next row number in Excel
            next_row = excel_row_count + 2  # +2 because row 1 is the header
            
            synced_count = 0
            
            for form_row in new_rows:
                # Generate Registration ID
                reg_id = f"BYOD-{datetime.now().year}{next_row - 1:06d}"
                
                # Map Google Form columns to Excel columns
                # Adjust indices based on your form structure
                excel_row_data = [
                     reg_id,                          # A: Registration ID (auto-generated)
                    form_row[0] if len(form_row) > 0 else '',   # B: Timestamp
                    form_row[1] if len(form_row) > 1 else '',   # C: Name (Full Name from form)
                    form_row[4] if len(form_row) > 4 else '',   # D: Department
                    '',                                          # E: Classification (NOT IN FORM - empty)
                    form_row[7] if len(form_row) > 7 else '',   # F: Device Type
                    form_row[8] if len(form_row) > 8 else '',   # G: Device Make/Model
                    form_row[9] if len(form_row) > 9 else '',   # H: Operating System
                    form_row[10] if len(form_row) > 10 else '', # I: MAC Address
                    form_row[11] if len(form_row) > 11 else '', # J: Serial Number
                    form_row[2] if len(form_row) > 2 else '',   # K: Email Address
                    form_row[3] if len(form_row) > 3 else '',   # L: Phone Number
                    form_row[5] if len(form_row) > 5 else '',   # M: Supervisor Name
                    form_row[6] if len(form_row) > 6 else '',   # N: Supervisor Email
                    'Pending',                                   # O: Status (auto)
                    '',                                          # P: Approved By (filled later)
                    '',                                          # Q: Approval Date (filled later)
                    ''                                           # R: Admin Remarks (filled later)
                ]
                
                # Write to Excel
                for col, value in enumerate(excel_row_data, 1):
                    sheet.cell(row=next_row, column=col).value = value
                
                print(f"  ✓ Synced: {excel_row_data[2]} - {reg_id}")
                next_row += 1
                synced_count += 1
            
            # Save Excel
            wb.save(self.excel_file)
            return synced_count
            
        except Exception as e:
            print(f"Error during sync: {e}")
            import traceback
            traceback.print_exc()
            return 0
    
    def run_continuous_sync(self, interval=60):
        """
        Run continuous sync
        
        Args:
            interval: Seconds between each sync check (default: 60)
        """
        print("\n" + "="*70)
        print("GOOGLE SHEETS AUTO-SYNC STARTED")
        print("="*70)
        print(f"Checking for new responses every {interval} seconds")
        print("Press Ctrl+C to stop")
        print("="*70 + "\n")
        
        try:
            while True:
                print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Checking for new responses...")
                
                new_count = self.sync_new_responses()
                
                if new_count > 0:
                    print(f"✓ Synced {new_count} new response(s)")
                    print("Running automation...")
                    
                    # Run the BYOD automation
                    import subprocess
                    print("\nRunning BYOD Automation...")
                    try:
                        # Import and run directly (no subprocess issues!)
                        from byod_automation import BYODAutomation

                        automation = BYODAutomation('NITDA_BYOD_Database.xlsx')
                        automation.run_automation()

                        print("✓ Automation completed\n")
                    except Exception as e:
                        print(f"✗ Error: {e}\n")
                
                else:
                    print("  No new responses")
                
                # Wait before next check
                time.sleep(interval)
                
        except KeyboardInterrupt:
            print("\n\nAuto-sync stopped by user")
        except Exception as e:
            print(f"\n\nError: {e}")
            import traceback
            traceback.print_exc()


def setup_google_sheets_api():
    """Guide user through Google Sheets API setup"""
    print("\n" + "="*70)
    print("GOOGLE SHEETS API SETUP GUIDE")
    print("="*70)
    
    print("""
This script requires a Google Service Account to access Google Sheets.

STEP 1: Create Google Cloud Project
------------------------------------
1. Go to: https://console.cloud.google.com
2. Click "Select a project" → "New Project"
3. Name: "NITDA BYOD System"
4. Click "Create"

STEP 2: Enable Google Sheets API
---------------------------------
1. In the search bar, type "Google Sheets API"
2. Click "Enable"
3. Also search for "Google Drive API" and enable it

STEP 3: Create Service Account
-------------------------------
1. Go to "IAM & Admin" → "Service Accounts"
2. Click "Create Service Account"
3. Name: "BYOD Automation"
4. Click "Create and Continue"
5. Skip optional steps, click "Done"

STEP 4: Create Key
------------------
1. Click on the service account you just created
2. Go to "Keys" tab
3. Click "Add Key" → "Create new key"
4. Choose "JSON"
5. Click "Create"
6. Save the file as: credentials.json (in your project folder)

STEP 5: Share Google Sheet
---------------------------
1. Open the downloaded credentials.json file
2. Find the "client_email" field (looks like: xxx@xxx.iam.gserviceaccount.com)
3. Copy this email address
4. Open your Google Sheet (the one linked to your form)
5. Click "Share"
6. Paste the service account email
7. Give it "Editor" access
8. Click "Send"

STEP 6: Get Your Spreadsheet Name
----------------------------------
1. Open your Google Sheet
2. Copy the exact name from the top (e.g., "NITDA BYOD Registration Database")

You're ready to run the auto-sync!
""")


# Simple CSV-based sync (no API required)
def simple_csv_sync(csv_file='form_responses.csv', excel_file='NITDA_BYOD_Database.xlsx'):
    """
    Simple sync from CSV export (no API needed)
    Download Google Sheet as CSV and run this
    """
    import csv
    
    if not os.path.exists(csv_file):
        print(f"❌ {csv_file} not found!")
        print("Download your Google Sheet as CSV first:")
        print("File → Download → Comma Separated Values (.csv)")
        return
    
    # Read CSV
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        headers = next(reader)
        csv_data = list(reader)
    
    # Get existing Excel row count
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb['Device Registrations']
    
    existing_count = 0
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value:
            existing_count += 1
    
    # Get only new rows
    new_rows = csv_data[existing_count:]
    
    if not new_rows:
        print("No new responses to sync")
        return
    
    next_row = existing_count + 2
    
    for form_row in new_rows:
        reg_id = f"BYOD-{datetime.now().year}{next_row - 1:06d}"
        
        # Map columns
        sheet.cell(row=next_row, column=1).value = reg_id
        sheet.cell(row=next_row, column=2).value = form_row[0] if len(form_row) > 0 else ''   # Timestamp
        sheet.cell(row=next_row, column=3).value = form_row[2] if len(form_row) > 2 else ''   # Name
        sheet.cell(row=next_row, column=4).value = form_row[4] if len(form_row) > 4 else ''   # Department
        sheet.cell(row=next_row, column=5).value = ''                                          # Classification (empty)
        sheet.cell(row=next_row, column=6).value = form_row[7] if len(form_row) > 7 else ''   # Device Type
        sheet.cell(row=next_row, column=7).value = form_row[8] if len(form_row) > 8 else ''   # Device Model
        sheet.cell(row=next_row, column=8).value = form_row[9] if len(form_row) > 9 else ''   # OS
        sheet.cell(row=next_row, column=9).value = form_row[10] if len(form_row) > 10 else '' # MAC
        sheet.cell(row=next_row, column=10).value = form_row[11] if len(form_row) > 11 else '' # Serial
        sheet.cell(row=next_row, column=11).value = form_row[1] if len(form_row) > 1 else ''  # Email
        sheet.cell(row=next_row, column=12).value = form_row[3] if len(form_row) > 3 else ''  # Phone
        sheet.cell(row=next_row, column=13).value = form_row[5] if len(form_row) > 5 else ''  # Supervisor Name
        sheet.cell(row=next_row, column=14).value = form_row[6] if len(form_row) > 6 else ''  # Supervisor Email
        sheet.cell(row=next_row, column=15).value = 'Pending'
                
        print(f"✓ Added: {form_row[2] if len(form_row) > 2 else 'Unknown'}")
        next_row += 1
    
    wb.save(excel_file)
    print(f"\n✓ Synced {len(new_rows)} new responses!")
    
    # Run automation
    import subprocess
    subprocess.run(['python', 'byod_automation.py'])


if __name__ == "__main__":
    import sys
    
    print("\n" + "="*70)
    print("GOOGLE SHEETS SYNC OPTIONS")
    print("="*70)
    
    print("\n1. Auto-sync with Google Sheets API (Recommended - fully automated)")
    print("2. Manual CSV sync (Simple - no API setup)")
    print("3. Show API setup instructions")
    
    choice = input("\nEnter choice (1/2/3): ").strip()
    
    if choice == "1":
        # Check if credentials exist
        if not os.path.exists('credentials.json'):
            print("\n❌ credentials.json not found!")
            print("\nPlease complete Google Sheets API setup first.")
            print("Run this script and choose option 3 for instructions.")
            sys.exit(1)
        
        spreadsheet_name = input("\nEnter your Google Sheet name: ").strip()
        
        sync = GoogleSheetsAutoSync(
            credentials_file='credentials.json',
            spreadsheet_name=spreadsheet_name,
            excel_file='NITDA_BYOD_Database.xlsx'
        )
        
        sync.run_continuous_sync(interval=60)  # Check every 60 seconds
        
    elif choice == "2":
        print("\nManual CSV Sync")
        print("1. Open your Google Sheet")
        print("2. File → Download → CSV")
        print("3. Save as 'form_responses.csv'")
        input("\nPress Enter when ready...")
        
        simple_csv_sync()
        
    elif choice == "3":
        setup_google_sheets_api()
    
    else:
        print("Invalid choice")
