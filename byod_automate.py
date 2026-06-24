import os
import sys
import time
import random
import uuid
import re
from datetime import datetime
from pathlib import Path
import openpyxl
import gspread
from google.oauth2.service_account import Credentials
from dotenv import load_dotenv
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# Load environment variables from .env file (SECURITY: P0 - Externalize credentials)
load_dotenv()

# Import your existing BYOD core logic directly
from byod_automation import BYODAutomation

# ==========================================
# INPUT VALIDATION & SANITIZATION (SECURITY: P0 - Prevent Formula Injection & Data Corruption)
# ==========================================
def sanitize_excel_input(value: str, max_length: int = 255) -> str:
    """
    Escapes Excel formula injection and enforces length limits.
    
    SECURITY FIX: Prevents attackers from injecting formulas like:
    =cmd|'/c powershell -Command "Get-Process"'
    
    - Prefixes dangerous characters (=, +, -, @, \t, \r) with single quote
    - Enforces max length to prevent buffer overflows
    - Returns empty string if value is None/invalid
    """
    if not isinstance(value, str):
        return ""
    
    value = value.strip()[:max_length]  # Trim to max length
    
    # Escape Excel formula injection attempts
    if value and value[0] in ('=', '+', '-', '@', '\t', '\r'):
        value = "'" + value
    
    return value

def validate_mac_address(mac: str) -> bool:
    """Validates MAC address format (XX:XX:XX:XX:XX:XX or XX-XX-XX-XX-XX-XX)."""
    if not mac:
        return False
    mac_pattern = r'^([0-9A-Fa-f]{2}[:-]){5}([0-9A-Fa-f]{2})$'
    return bool(re.match(mac_pattern, mac))

def validate_email(email: str) -> bool:
    """Basic email format validation."""
    if not email:
        return False
    email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return bool(re.match(email_pattern, email))

def validate_phone(phone: str) -> bool:
    """Basic phone format validation (allows digits, spaces, hyphens, parentheses)."""
    if not phone:
        return False
    phone_pattern = r'^[+]?[\d\s\-()]{10,}$'  # At least 10 characters of phone-like content
    return bool(re.match(phone_pattern, phone))

def safe_get_google_row_value(google_row: list, index: int, default: str = "") -> str:
    """
    Safely retrieves a value from Google Forms data with bounds checking.
    
    SECURITY FIX: Prevents IndexError crashes if Google Forms schema changes.
    Returns empty string if index is out of bounds.
    """
    if 0 <= index < len(google_row):
        value = google_row[index]
        return sanitize_excel_input(str(value) if value else "")
    return default

# ==========================================
# CONCURRENCY-SAFE FILE WRAPPERS
# ==========================================
def safe_load_workbook(filename, max_retries=5, initial_delay=0.5):
    """Safely loads a workbook, retrying if locked by a human admin."""
    delay = initial_delay
    for attempt in range(max_retries):
        try:
            return openpyxl.load_workbook(filename)
        except (PermissionError, IOError):
            if attempt == max_retries - 1:
                raise PermissionError(f"❌ Could not read {filename}. File is locked by another program.")
            sleep_time = delay + random.uniform(0.1, 0.4)
            time.sleep(sleep_time)
            delay *= 2

def safe_save_workbook(wb, filename, max_retries=5, initial_delay=0.5):
    """Safely saves a workbook, retrying if locked by an open instance of Excel."""
    delay = initial_delay
    for attempt in range(max_retries):
        try:
            wb.save(filename)
            return True
        except (PermissionError, IOError):
            if attempt == max_retries - 1:
                print(f"❌ CRITICAL ERROR: Unable to save changes to {filename}. Data may be lost!")
                return False
            sleep_time = delay + random.uniform(0.1, 0.4)
            print(f"⚠️ Excel file locked by an administrator. Retrying save in {sleep_time:.1f}s...")
            time.sleep(sleep_time)
            delay *= 2

# ==========================================
# UNIFIED ORCHESTRATION ENGINE
# ==========================================
class BYODUnifiedEngine:
    def __init__(self, excel_file=None, credentials_file=None, google_sheet_name=None, check_interval=None):
        """
        SECURITY: P0 - Load configuration from environment variables instead of hardcoding.
        Falls back to parameters for backward compatibility.
        """
        # Load from .env file, fallback to provided parameters, then defaults
        self.excel_file = excel_file or os.getenv('DATABASE_NAME', 'NITDA_BYOD_Database.xlsx')
        self.credentials_file = credentials_file or os.getenv('CREDENTIALS_JSON', 'credentials.json')
        self.google_sheet_name = google_sheet_name or os.getenv('GOOGLE_SHEET_NAME', '')
        self.check_interval = check_interval or int(os.getenv('CHECK_INTERVAL', '30'))
        
        # Track which rows we've already synced from Google Sheets (more reliable than row count)
        self.synced_timestamp_hashes = set()
        
        # File system watcher for detecting manual edits (replaces timestamp polling)
        self.observer = None
        self.file_event_detected = False
        
        # Authenticate Google Sheets once at initialization
        self.gc = None
        self.setup_google_api()
        
        print("="*70)
        print("⚙️  CONFIGURATION LOADED")
        print("="*70)
        print(f"Database File  : {self.excel_file}")
        print(f"Google Sheet   : {self.google_sheet_name}")
        print(f"Check Interval : {self.check_interval}s")
        print()

    def setup_google_api(self):
        """Initializes connection to Google Sheets API with proper error handling."""
        if not os.path.exists(self.credentials_file):
            print(f"⚠️  Warning: '{self.credentials_file}' not found. Google Form Sync disabled.")
            print(f"   Create credentials at: https://console.cloud.google.com/apis/")
            return
        try:
            scopes = [
                'https://www.googleapis.com/auth/spreadsheets',
                'https://www.googleapis.com/auth/drive'
            ]
            creds = Credentials.from_service_account_file(self.credentials_file, scopes=scopes)
            self.gc = gspread.authorize(creds)
            print("✅ Google Sheets API authenticated successfully.")
        except FileNotFoundError:
            print(f"❌ Credentials file '{self.credentials_file}' is invalid JSON.")
        except Exception as e:
            print(f"❌ Failed to initialize Google Sheets: {type(e).__name__}: {str(e)[:100]}")
            print(f"   Ensure credentials.json has the correct format and permissions.")

    def get_file_modified_time(self):
        """Fetches file modification timestamp safely."""
        try:
            return os.path.getmtime(self.excel_file)
        except:
            return 0

    # ==========================================
    # FILE SYSTEM EVENT HANDLER (SECURITY: P1 - Replace polling with events)
    # ==========================================
    class ExcelFileWatcher(FileSystemEventHandler):
        """Detects when the Excel file is modified by a human administrator."""
        def __init__(self, engine):
            self.engine = engine
        
        def on_modified(self, event):
            """Called when the Excel file is modified."""
            if event.src_path.endswith(self.engine.excel_file) or event.src_path.endswith('.xlsx'):
                # Debounce: Multiple events fire for one save operation
                if not self.engine.file_event_detected:
                    self.engine.file_event_detected = True
                    print(f"🔍 File modification detected: {Path(event.src_path).name}")
    
    def start_file_watcher(self):
        """Starts monitoring the Excel file for manual edits using file system events."""
        try:
            file_directory = str(Path(self.excel_file).parent.absolute())
            self.observer = Observer()
            self.observer.schedule(self.ExcelFileWatcher(self), file_directory, recursive=False)
            self.observer.start()
            print(f"✅ File watcher started for: {file_directory}")
        except Exception as e:
            print(f"⚠️  Could not start file watcher: {e}. Falling back to polling.")
            self.observer = None
    
    def stop_file_watcher(self):
        """Stops the file system event observer."""
        if self.observer:
            self.observer.stop()
            self.observer.join()
            print("🛑 File watcher stopped.")

    def sync_google_forms(self) -> bool:
        """
        Connects to Google Sheets, compares entries, and syncs new data to Excel.
        
        SECURITY IMPROVEMENTS:
        - P0: Sanitizes all external input to prevent formula injection
        - P0: Validates data types and formats before writing to Excel
        - P1: Uses UUID for registration IDs (no duplicate risk)
        - P1: Bounds-checks array access to handle schema changes gracefully
        
        Returns True if new entries were written, False otherwise.
        """
        if not self.gc:
            return False
            
        try:
            # 1. Fetch live remote data from Google Sheets
            sheet = self.gc.open(self.google_sheet_name).sheet1
            google_rows = sheet.get_all_values()
            if len(google_rows) <= 1:  # Only header row exists
                return False
                
            # 2. Open local Excel to prepare for appending
            wb = safe_load_workbook(self.excel_file)
            ws = wb['Device Registrations']
            
            # 3. Compare entries using timestamp+email hash (more robust than row count)
            new_entries_to_add = []
            for google_row in google_rows[1:]:  # Skip header
                # Create a unique hash of this entry to track syncing
                timestamp = safe_get_google_row_value(google_row, 0)
                email = safe_get_google_row_value(google_row, 2)
                entry_hash = hash(f"{timestamp}:{email}")
                
                if entry_hash not in self.synced_timestamp_hashes:
                    new_entries_to_add.append(google_row)
                    self.synced_timestamp_hashes.add(entry_hash)
            
            if not new_entries_to_add:
                return False
            
            print(f"🟢 Found {len(new_entries_to_add)} new form submission(s). Syncing with validation...")
            
            # 4. Append new rows with comprehensive validation
            for google_row in new_entries_to_add:
                next_excel_row = ws.max_row + 1
                
                # Generate UUID-based registration ID (SECURITY: P0 - Guarantee uniqueness)
                reg_id = f"BYOD-{uuid.uuid4().hex[:12].upper()}"
                
                # Extract and validate all fields with bounds checking and sanitization
                timestamp = safe_get_google_row_value(google_row, 0)
                name = safe_get_google_row_value(google_row, 1)
                email = safe_get_google_row_value(google_row, 2)
                department = safe_get_google_row_value(google_row, 4)
                phone = safe_get_google_row_value(google_row, 3)
                supervisor_name = safe_get_google_row_value(google_row, 5)
                supervisor_email = safe_get_google_row_value(google_row, 6)
                device_type = safe_get_google_row_value(google_row, 7)
                make_model = safe_get_google_row_value(google_row, 8)
                os_type = safe_get_google_row_value(google_row, 9)
                mac_address = safe_get_google_row_value(google_row, 10)
                serial_number = safe_get_google_row_value(google_row, 11)
                
                # SECURITY: P1 - Validate critical fields
                validation_errors = []
                if email and not validate_email(email):
                    validation_errors.append(f"Invalid email format: {email}")
                if phone and not validate_phone(phone):
                    validation_errors.append(f"Invalid phone format: {phone}")
                if mac_address and not validate_mac_address(mac_address):
                    validation_errors.append(f"Invalid MAC address format: {mac_address}")
                
                if validation_errors:
                    print(f"   ⚠️  Row {next_excel_row}: Validation warnings - {'; '.join(validation_errors)}")
                
                # Write sanitized data to Excel (formula injection already prevented by sanitize_excel_input)
                ws.cell(row=next_excel_row, column=1, value=reg_id)
                ws.cell(row=next_excel_row, column=2, value=timestamp)
                ws.cell(row=next_excel_row, column=3, value=name)
                ws.cell(row=next_excel_row, column=4, value=department)
                ws.cell(row=next_excel_row, column=5, value="")  # Reserved for future use
                ws.cell(row=next_excel_row, column=6, value=device_type)
                ws.cell(row=next_excel_row, column=7, value=make_model)
                ws.cell(row=next_excel_row, column=8, value=os_type)
                ws.cell(row=next_excel_row, column=9, value=mac_address)
                ws.cell(row=next_excel_row, column=10, value=serial_number)
                ws.cell(row=next_excel_row, column=11, value=email)
                ws.cell(row=next_excel_row, column=12, value=phone)
                ws.cell(row=next_excel_row, column=13, value=supervisor_name)
                ws.cell(row=next_excel_row, column=14, value=supervisor_email)
                ws.cell(row=next_excel_row, column=15, value="Pending")
                
                print(f"   ✅ Synced: {name} ({reg_id})")
            
            # Save data safely
            if safe_save_workbook(wb, self.excel_file):
                return True
            return False
            
        except Exception as e:
            print(f"❌ Google Sheets sync error: {type(e).__name__}: {str(e)[:100]}")
            return False

    def check_manual_excel_edits(self) -> bool:
        """
        Checks if a human administrator manually updated the spreadsheet.
        Uses file system events instead of timestamp polling (SECURITY: P1 - More reliable).
        """
        if self.file_event_detected:
            print(f"✅ Manual edit confirmed. Running lifecycle management...")
            time.sleep(1)  # Brief settle time to ensure file is fully saved
            self.file_event_detected = False
            return True
        return False

    def run_engine(self):
        """
        Main continuous orchestration loop.
        Coordinates Google Forms sync, manual edits detection, and lifecycle management.
        """
        # Start file system watcher for detecting manual edits
        self.start_file_watcher()
        
        print("\n" + "="*70)
        print("🚀 NITDA BYOD UNIFIED AUTOMATION ENGINE - ACTIVE")
        print("="*70)
        print(f"Target Database : {self.excel_file}")
        print(f"Google Sheet    : {self.google_sheet_name}")
        print(f"Monitor Interval: {self.check_interval}s")
        print("\nEngine running. Press Ctrl+C to stop gracefully.\n")

        try:
            while True:
                # Flag to monitor if processing needs execution during this cycle
                should_run_automation = False
                
                # Action 1: Sync new form registrations from Google Sheets
                if self.sync_google_forms():
                    should_run_automation = True
                
                # Action 2: Check if local Excel was manually edited by an IT Officer
                if self.check_manual_excel_edits():
                    should_run_automation = True
                
                # Action 3: Execute lifecycle management logic if triggered
                if should_run_automation:
                    try:
                        print(f"\n▶️  Triggering Automation [{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]")
                        # Instantiate and execute core automation processing
                        automation_processor = BYODAutomation(self.excel_file)
                        automation_processor.run_automation()
                        print(f"✅ Automation cycle completed.\n")
                    except FileNotFoundError as e:
                        print(f"❌ File error: {e}")
                    except Exception as e:
                        print(f"❌ Automation error: {type(e).__name__}: {str(e)[:100]}")
                
                # Sleep until the next evaluation tick
                time.sleep(self.check_interval)
                
        except KeyboardInterrupt:
            print("\n\n🛑 Engine shutting down...")
            self.stop_file_watcher()
            print("👋 Unified Automation Engine stopped safely.\n")

# ==========================================
# BOOTSTRAP INITIALIZATION
# ==========================================
if __name__ == "__main__":
    """
    SECURITY: P0 - All configuration loaded from .env file, no hardcoding.
    For first-time setup, ensure .env file contains:
    - DATABASE_NAME: Path to Excel file
    - CREDENTIALS_JSON: Path to Google service account JSON
    - GOOGLE_SHEET_NAME: Name of the target Google Sheet
    - CHECK_INTERVAL: Seconds between checks (default: 30)
    """
    
    # Load environment variables
    load_dotenv()
    
    # Get configuration with validation
    database_name = os.getenv('DATABASE_NAME', '').strip()
    credentials_json = os.getenv('CREDENTIALS_JSON', '').strip()
    google_sheet_name = os.getenv('GOOGLE_SHEET_NAME', '').strip()
    check_interval = int(os.getenv('CHECK_INTERVAL', '30'))
    
    # Validate configuration
    if not database_name:
        print("❌ ERROR: DATABASE_NAME not set in .env file")
        sys.exit(1)
    if not credentials_json:
        print("⚠️  WARNING: CREDENTIALS_JSON not set. Google Sheets sync disabled.")
    if not google_sheet_name:
        print("❌ ERROR: GOOGLE_SHEET_NAME not set in .env file")
        sys.exit(1)
    
    # Check if files exist
    if not os.path.exists(database_name):
        print(f"❌ ERROR: Database file not found: {database_name}")
        sys.exit(1)
    
    # Initialize and run engine
    engine = BYODUnifiedEngine(
        excel_file=database_name,
        credentials_file=credentials_json,
        google_sheet_name=google_sheet_name,
        check_interval=check_interval
    )
    
    engine.run_engine()