"""
NITDA BYOD Automation System
This script automates the device registration workflow:
1. Reads new entries from Excel
2. Sends email notifications at each stage
3. Generates QR codes for approved devices
4. Schedules IT inspections
"""
import os
import time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl
import qrcode
import smtplib
import json

class BYODAutomation:
    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.wb = openpyxl.load_workbook(excel_file)
        self.registrations = self.wb['Device Registrations']
        self.inspections = self.wb['IT Inspection']
        self.settings = self.wb['Automation Settings']
        self.templates = self.wb['Email Templates']
        
        # Load email settings
        self.smtp_server = self.get_setting('SMTP Server')
        self.smtp_port = int(self.get_setting('SMTP Port'))
        self.sender_email = self.get_setting('Sender Email')
        self.sender_password = os.getenv('GMAIL_APP_PASSWORD', 'mvkd inaq hsyi kqnx')
        self.it_email = self.get_setting('IT Email')
        self.inspection_lead_time = int(self.get_setting('Inspection Lead Time (days)'))
        
        # Persistent SMTP connection
        self.smtp_connection = None
        
    def get_setting(self, setting_name):
        """Retrieve setting value from Automation Settings sheet"""
        for row in self.settings.iter_rows(min_row=3, values_only=False):
            if row[0].value == setting_name:
                return row[1].value
        return None
    
    def get_email_template(self, template_name):
        """Retrieve email template from Email Templates sheet"""
        for row in self.templates.iter_rows(min_row=4, values_only=False):
            if row[0].value == template_name:
                return {
                    'subject': row[1].value,
                    'body': row[2].value
                }
        return None
    
    def get_smtp_connection(self):
        """Get or create persistent SMTP connection with retry logic"""
        max_retries = 3
        retry_count = 0
        
        while retry_count < max_retries:
            try:
                # Create fresh connection if needed
                if self.smtp_connection is None:
                    if self.smtp_port == 465:
                        self.smtp_connection = smtplib.SMTP_SSL(
                            self.smtp_server, 
                            self.smtp_port, 
                            timeout=30
                        )
                    else:
                        self.smtp_connection = smtplib.SMTP(
                            self.smtp_server, 
                            self.smtp_port, 
                            timeout=30
                        )
                        self.smtp_connection.ehlo()
                        self.smtp_connection.starttls()
                        self.smtp_connection.ehlo()
                    
                    self.smtp_connection.login(self.sender_email, self.sender_password)
                    print(f"  \u2192 SMTP connection established")
                    return self.smtp_connection
                
                # Verify existing connection is alive with NOOP
                try:
                    self.smtp_connection.noop()
                    return self.smtp_connection
                except:
                    # Connection dead, reset it
                    self.smtp_connection = None
                    raise Exception("Connection lost, reconnecting...")
                    
            except Exception as e:
                print(f"  \u26a0 SMTP connection error (attempt {retry_count + 1}/{max_retries}): {e}")
                self.smtp_connection = None
                retry_count += 1
                
                if retry_count < max_retries:
                    time.sleep(2)  # Wait before retry
        
        raise Exception(f"Failed to connect to SMTP after {max_retries} attempts")
    
    def close_smtp_connection(self):
        """Properly close SMTP connection"""
        try:
            if self.smtp_connection:
                self.smtp_connection.quit()
                print("  \u2192 SMTP connection closed")
        except:
            pass
        finally:
            self.smtp_connection = None

    def send_email(self, to_email, subject, body, attachment=None):
        """Send email with optional attachment and retry logic"""
        try:
            msg = MIMEMultipart('alternative') if not attachment else MIMEMultipart()
            msg['From'] = self.sender_email
            msg['To'] = to_email
            msg['Subject'] = subject
            
            msg.attach(MIMEText(body, 'plain'))
            
            if attachment:
                msg.attach(attachment)
            
            # Get SMTP connection
            server = self.get_smtp_connection()
            server.send_message(msg)
            
            print(f"  \u2713 Email sent to {to_email}")
            return True
            
        except smtplib.SMTPException as e:
            print(f"  \u2717 SMTP error: {e}")
            self.smtp_connection = None  # Reset on error
            return False
        except Exception as e:
            print(f"  \u2717 Email error: {e}")
            return False
    
    def process_new_registrations(self):
        """Process new registrations and send confirmation emails"""
        processed_count = 0
        
        for row in self.registrations.iter_rows(min_row=2, values_only=False):
            reg_id = row[0].value
            if not reg_id:
                continue
            
            status = row[14].value  # Status column
            admin_remarks = row[17].value #admin remarks colum in the excel spreadsheet
            # Check if confirmation email needs to be sent
            if status == 'Pending' or not status and admin_remarks != 'Emails Sent':
                # Get template
                template = self.get_email_template('Registration Confirmation')
                
                # Format email
                email_body = template['body'].format(
                    name=row[2].value,
                    registration_id=reg_id,
                    device_model=row[6].value,
                    date=datetime.now().strftime('%Y-%m-%d  %H:%M:%S')
                )
                
                subject = template['subject'].format(registration_id=reg_id)
                
                # Send to user
                user_email = row[10].value
                if user_email:
                    self.send_email(user_email, subject, email_body)
                
                # Send supervisor approval request
                self.send_supervisor_approval(row)

                #Mark as emailed
                row[17].value = 'Emails Sent' #This fills up admin row so it can skip when auto sync and not send mail multiple times
                
                processed_count += 1
        
        return processed_count
 
      ###commenting out the previous supervisor approval function start
   # def send_supervisor_approval(self, row):
   #     """Send approval request to supervisor"""
    #    template = self.get_email_template('Supervisor Approval Request')
     #   
      #  supervisor_email = row[13].value
       # if not supervisor_email:
        #    return
        
        #email_body = template['body'].format(
         #   supervisor_name=row[12].value,
          ##  name=row[2].value,
            #department=row[3].value,
            #device_model=row[6].value,
            #registration_id=row[0].value
        #)
        
       # subject = template['subject'].format(name=row[2].value)
#        self.send_email(supervisor_email, subject, email_body)
   ###comment out supervisor approval function end 
   
   
#####################################################################new function start
    def send_supervisor_approval(self, row):
        """Send approval request to supervisor with approve/reject buttons"""
        supervisor_email = row[13].value
        if not supervisor_email:
            return
        
        reg_id = row[0].value
        name = row[2].value
        department = row[3].value
        device_model = row[6].value
        supervisor_name = row[12].value
        
        # Get approval server URL
        # The server URL here is currently the ngrok URL. Change to public server URL when deployed.
        server_url = "https://biostatical-penetratingly-gerri.ngrok-free.dev"
        approval_link = f"{server_url}/approve/{reg_id}"
        
        # Plain text version
        text_body = f"""Dear {supervisor_name},

A new BYOD registration requires your approval.

Intern's Name: {name}
Department: {department}
Device: {device_model}
Registration ID: {reg_id}

To approve or reject this device, please visit:
{approval_link}

Best regards,
NITDA BYOD System"""
        
        # HTML version with buttons
        html_body = f"""
<!DOCTYPE html>
<html>
<head>
    <style>
        body {{
            font-family: Arial, sans-serif;
            line-height: 1.6;
            color: #333;
        }}
        .container {{
            max-width: 600px;
            margin: 0 auto;
            padding: 20px;
        }}
        .header {{
            background: #00A86B;
            color: white;
            padding: 20px;
            text-align: center;
            border-radius: 8px 8px 0 0;
        }}
        .content {{
            background: #f9f9f9;
            padding: 30px;
            border: 1px solid #ddd;
            border-top: none;
        }}
        .info-box {{
            background: white;
            padding: 15px;
            border-left: 4px solid #00A86B;
            margin: 20px 0;
        }}
        .info-row {{
            padding: 8px 0;
            border-bottom: 1px solid #eee;
        }}
        .info-row:last-child {{
            border-bottom: none;
        }}
        .label {{
            font-weight: bold;
            color: #666;
            display: inline-block;
            width: 150px;
        }}
        .button-container {{
            text-align: center;
            margin: 30px 0;
        }}
        .button {{
            display: inline-block;
            padding: 15px 40px;
            margin: 10px;
            text-decoration: none;
            border-radius: 8px;
            font-weight: bold;
            font-size: 16px;
            background: #00A86B;
            color: white !important;
        }}
        .button:hover {{
            background: #007850;
        }}
        .footer {{
            text-align: center;
            padding: 20px;
            color: #666;
            font-size: 12px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>BYOD Approval Required</h1>
        </div>
        <div class="content">
            <p>Dear {supervisor_name},</p>
            <p>A new BYOD device registration requires your endorsement.</p>
            
            <div class="info-box">
                <div class="info-row">
                    <span class="label">Registration ID:</span>
                    <span>{reg_id}</span>
                </div>
                <div class="info-row">
                    <span class="label">Intern's Name:</span>
                    <span>{name}</span>
                </div>
                <div class="info-row">
                    <span class="label">Department:</span>
                    <span>{department}</span>
                </div>
                <div class="info-row">
                    <span class="label">Device:</span>
                    <span>{device_model}</span>
                </div>
            </div>
            
            <div class="button-container">
                <a href="{approval_link}" class="button">
                    ✓ REVIEW & APPROVE/REJECT
                </a>
            </div>
            
            <p style="text-align: center; color: #666; font-size: 14px;">
                Click the button above to review and approve/reject this registration.
            </p>
        </div>
        <div class="footer">
            <p>This is an automated email from NITDA BYOD Management System</p>
        </div>
    </div>
</body>
</html>
"""
        
        # Create message with both plain text and HTML
        try:
            msg = MIMEMultipart('alternative')
            msg['From'] = self.sender_email
            msg['To'] = supervisor_email
            msg['Subject'] = f'BYOD Approval Required - {name}'
            
            part1 = MIMEText(text_body, 'plain')
            part2 = MIMEText(html_body, 'html')
            msg.attach(part1)
            msg.attach(part2)
            
            server = self.get_smtp_connection()
            server.send_message(msg)
            
            print(f"  \u2713 Approval email sent to {supervisor_email}")
            return True
        except Exception as e:
            print(f"  \u2717 Approval email error: {e}")
            self.smtp_connection = None
            return False


    #new function end#####################################################################
    
   
    def process_approved_devices(self):
        """Process approved devices and schedule IT inspections"""
        processed_count = 0
        
        if not hasattr(self, 'inspections') or self.inspections is None:
            print("  \u2717 Inspections sheet not initialized")
            return 0
        
        for row in self.registrations.iter_rows(min_row=2, values_only=False):
            reg_id = row[0].value
            if not reg_id:
                continue
            
            status = row[14].value
            
            if status == 'Approved':
                # Check if IT inspection already scheduled
                inspection_exists = self.check_inspection_exists(reg_id)
                
                if not inspection_exists:
                    # Schedule inspection
                    inspection_date = datetime.now() + timedelta(days=self.inspection_lead_time)
                    
                    # Add to IT Inspection sheet
                    self.add_inspection_record(row, inspection_date)
                    
                    template = self.get_email_template('IT Inspection Schedule')
                    if template:
                        email_body = template['body'].format(
                            name=row[2].value,
                            registration_id=reg_id,
                            device_model=row[6].value,
                            inspection_date=inspection_date.strftime('%Y-%m-%d')
                        )
                        
                        user_email = row[10].value
                        if user_email:
                            subject = template['subject'].format(registration_id=reg_id)
                            self.send_email(user_email, subject, email_body)
                    
                    it_notification = f"""New device scheduled for inspection:

Registration ID: {reg_id}
Name: {row[2].value}
Device: {row[6].value}
Scheduled: {inspection_date.strftime('%Y-%m-%d')}

Contact: {row[10].value}"""
                    
                    self.send_email(self.it_email, f'New Inspection - {reg_id}', it_notification)
                    
                    processed_count += 1
        
        return processed_count
    
    def check_inspection_exists(self, reg_id):
        """Check if inspection record exists for registration ID"""
        for row in self.inspections.iter_rows(min_row=2, values_only=True):
            if row[0] == reg_id:
                return True
        return False
    
    def add_inspection_record(self, reg_row, inspection_date):
        """Add new inspection record"""
        next_row = self.inspections.max_row + 1
        inspection_id = f"INS-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        
        self.inspections.cell(row=next_row, column=1).value = reg_row[0].value
        self.inspections.cell(row=next_row, column=2).value = reg_row[2].value
        self.inspections.cell(row=next_row, column=3).value = reg_row[6].value
        self.inspections.cell(row=next_row, column=4).value = reg_row[9].value
        self.inspections.cell(row=next_row, column=5).value = inspection_id
        self.inspections.cell(row=next_row, column=6).value = inspection_date.strftime('%Y-%m-%d')
        self.inspections.cell(row=next_row, column=7).value = ''
        
        for col in range(8, 14):
            self.inspections.cell(row=next_row, column=col).value = ''
        
        self.inspections.cell(row=next_row, column=14).value = 'Pending'
        self.inspections.cell(row=next_row, column=15).value = ''
        self.inspections.cell(row=next_row, column=16).value = 'No'
        self.inspections.cell(row=next_row, column=17).value = ''
        
        print(f"  \u2192 Created IT Inspection: {inspection_id}")

    def generate_qr_code(self, reg_id, device_info):
        """Generate QR code for device pass"""
        # Create QR code data
        qr_data = {
            'registration_id': reg_id,
            'name': device_info['name'],
            'device_model': device_info['device_model'],
            'serial_number': device_info['serial_number'],
            'mac_address': device_info['mac_address'],
            'department': device_info['department'],
            'compliance_status': 'COMPLIANT',
            'issued_date': datetime.now().strftime('%Y-%m-%d'),
            'valid': True
        }
        
        # Generate QR code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_H,
            box_size=10,
            border=4,
        )
        qr.add_data(json.dumps(qr_data))
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Save to BytesIO for email attachment
        img_buffer = BytesIO()
        img.save(img_buffer, format='PNG')
        img_buffer.seek(0)
        
        return img_buffer
    
    def process_compliant_devices(self):
        """Process compliant devices and generate QR codes"""
        processed_count = 0
        
        for row in self.inspections.iter_rows(min_row=2, values_only=False):
            reg_id = row[0].value
            if not reg_id:
                continue
            
            compliance_status = row[13].value  # Compliance Status
            qr_generated = row[15].value  # QR Code Generated
            
            if compliance_status == 'Compliant' and qr_generated != 'Yes':
                # Get device info from registrations
                device_info = self.get_device_info(reg_id)
                
                if device_info:
                    # Generate QR code
                    qr_image = self.generate_qr_code(reg_id, device_info)
                    
                    # Create email attachment
                    qr_attachment = MIMEImage(qr_image.read())
                    qr_attachment.add_header('Content-Disposition', 'attachment', 
                                           filename=f'{reg_id}_pass.png')
                    
                    # Send email with QR code
                    template = self.get_email_template('QR Code Pass Issued')
                    email_body = template['body'].format(
                        name=device_info['name'],
                        registration_id=reg_id,
                        device_model=device_info['device_model']
                    )
                    
                    subject = template['subject'].format(registration_id=reg_id)
                    
                    self.send_email(device_info['email'], subject, email_body, qr_attachment)
                    
                    # Update inspection record
                    row[15].value = 'Yes'
                    row[16].value = datetime.now().strftime('%Y-%m-%d')
                    
                    # Save QR code to file
                    qr_image.seek(0)
                    os.makedirs('qr_codes', exist_ok=True)
                    with open(f'qr_codes/{reg_id}_pass.png', 'wb') as f:
                        f.write(qr_image.read())
                    
                    processed_count += 1
        
        return processed_count
    
    def get_device_info(self, reg_id):
        """Get device information from registrations sheet"""
        for row in self.registrations.iter_rows(min_row=2, values_only=True):
            if row[0] == reg_id:
                return {
                    'name': row[2],
                    'department': row[3],
                    'device_model': row[6],
                    'mac_address': row[8],
                    'serial_number': row[9],
                    'email': row[10]
                }
        return None
    
    def save_changes(self):
        """Save all changes to Excel file"""
        self.wb.save(self.excel_file)
        print("  \u2192 Changes saved")
    
    def run_automation(self):
        """Run all automation processes"""
        print("=" * 60)
        print("NITDA BYOD Automation - Running")
        print("=" * 60)
        
        try:
            print("\n1. Processing new registrations...")
            new_regs = self.process_new_registrations()
            print(f"   ✓ Processed {new_regs} registrations\n")
            
            print("2. Processing approved devices...")
            approved = self.process_approved_devices()
            print(f"   ✓ Scheduled {approved} inspections\n")
            
            print("3. Generating QR codes...")
            qr_codes = self.process_compliant_devices()
            print(f"   ✓ Generated {qr_codes} QR codes\n")
            
            print("4. Saving changes...")
            self.save_changes()
            
        finally:
            self.close_smtp_connection()
        
        print("\n" + "=" * 60)
        print("✓ Automation completed successfully!")
        print("=" * 60)


# Main execution
if __name__ == "__main__":
    automation = BYODAutomation('NITDA_BYOD_Database.xlsx')
    
    try:
        automation.run_automation()
    finally:
        automation.close_smtp_connection()
