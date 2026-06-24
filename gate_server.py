"""
NITDA BYOD — Security Gate Server
Handles QR code scan requests from the gate_scanner.html terminal.

SECURITY FEATURES (P0/P1):
- API key authentication
- Excel formula injection prevention
- Rate limiting per IP
- Concurrent file access handling
- Request validation & sanitization
- Restricted CORS (LAN only)
- Comprehensive audit logging

Install: pip install flask openpyxl flask-cors python-dotenv
"""

from flask import Flask, request, jsonify
from flask_cors import CORS
from functools import wraps
from dotenv import load_dotenv
import openpyxl
import re
import os
import time
from datetime import datetime, timedelta
from collections import defaultdict
import threading

load_dotenv()
app = Flask(__name__)

# ── CONFIG ──────────────────────────────────────────────────────────────────
EXCEL_FILE = os.getenv('DATABASE_NAME', 'NITDA_BYOD_Database.xlsx')
GATE_API_KEY = os.getenv('GATE_API_KEY', '')  # SECURITY: P0 - Load from .env
ALLOWED_IPS = ['127.0.0.1', '::1']  # Add security post PC's LAN IP

RATE_LIMIT_REQUESTS = 60
RATE_LIMIT_WINDOW = 60
request_tracker = defaultdict(list)
tracker_lock = threading.Lock()

# CORS: restrict to localhost/LAN only (SECURITY: P1)
CORS(app, resources={
    r"/gate/*": {
        "origins": [
            "http://localhost:*",
            "http://127.0.0.1:*",
            "http://192.168.*",
            "http://10.*"
        ],
        "methods": ["GET", "POST"],
        "allow_headers": ["Content-Type", "X-API-Key"]
    }
})

# ── SECURITY HELPERS ────────────────────────────────────────────────────────

def sanitize_excel_input(value: str, max_length: int = 255) -> str:
    """
    SECURITY: P0 - Prevent Excel formula injection.
    Escapes dangerous characters that could execute formulas.
    """
    if not isinstance(value, str):
        return ""
    
    value = str(value).strip()[:max_length]
    
    # Escape formula injection attempts (=, +, -, @, \t, \r)
    if value and value[0] in ('=', '+', '-', '@', '\t', '\r'):
        value = "'" + value
    
    return value


def validate_input(value: str, pattern: str = None, max_length: int = 255) -> bool:
    """Validate input format and length."""
    if not value or not isinstance(value, str):
        return False
    if len(value) > max_length:
        return False
    if pattern and not re.match(pattern, value):
        return False
    return True


def check_rate_limit(ip: str) -> bool:
    """SECURITY: P1 - Rate limiting to prevent DDoS."""
    with tracker_lock:
        now = time.time()
        # Remove old entries
        request_tracker[ip] = [t for t in request_tracker[ip] 
                               if now - t < RATE_LIMIT_WINDOW]
        
        if len(request_tracker[ip]) >= RATE_LIMIT_REQUESTS:
            return False
        
        request_tracker[ip].append(now)
        return True


def require_api_key(f):
    """SECURITY: P0 - Require API key for gate endpoints."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not GATE_API_KEY:
            app.logger.warning("⚠️  GATE_API_KEY not set in .env - authentication disabled!")
            return f(*args, **kwargs)
        
        api_key = request.headers.get('X-API-Key', '')
        if api_key != GATE_API_KEY:
            return jsonify({
                'ok': False,
                'reason': 'UNAUTHORISED',
                'message': 'Invalid or missing API key'
            }), 401
        
        return f(*args, **kwargs)
    return decorated


def require_ip_whitelist(f):
    """SECURITY: P0 - IP whitelist check."""
    @wraps(f)
    def decorated(*args, **kwargs):
        # Get real IP (handles reverse proxy with X-Forwarded-For)
        client_ip = request.headers.get('X-Forwarded-For', request.remote_addr)
        if ',' in client_ip:
            client_ip = client_ip.split(',')[0].strip()
        
        # Check whitelist
        if ALLOWED_IPS and client_ip not in ALLOWED_IPS:
            app.logger.warning(f"[SECURITY] Blocked scan from unauthorized IP: {client_ip}")
            return jsonify({
                'ok': False,
                'reason': 'UNAUTHORISED_IP',
                'message': f'IP {client_ip} not authorized'
            }), 403
        
        return f(*args, **kwargs)
    return decorated


# ── EXCEL HELPERS ───────────────────────────────────────────────────────────

def safe_load_workbook(max_retries=5, initial_delay=0.2):
    """Safely load workbook with retry logic for concurrent access."""
    delay = initial_delay
    for attempt in range(max_retries):
        try:
            return openpyxl.load_workbook(EXCEL_FILE)
        except (PermissionError, IOError):
            if attempt == max_retries - 1:
                raise
            time.sleep(delay)
            delay *= 1.5


def safe_save_workbook(wb, max_retries=5, initial_delay=0.2):
    """Safely save workbook with retry logic."""
    delay = initial_delay
    for attempt in range(max_retries):
        try:
            wb.save(EXCEL_FILE)
            return True
        except (PermissionError, IOError):
            if attempt == max_retries - 1:
                return False
            time.sleep(delay)
            delay *= 1.5


def get_device_record(reg_id: str) -> dict | None:
    """
    Look up registration ID in Device Registrations sheet.
    Returns sanitized device info or None if not found.
    """
    if not validate_input(reg_id, r'^[A-Z0-9\-]{6,50}$', 50):
        app.logger.warning(f"[SECURITY] Invalid reg_id format: {reg_id}")
        return None
    
    try:
        wb = safe_load_workbook()
        sheet = wb['Device Registrations']
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and str(row[0]).strip() == reg_id.strip():
                device = {
                    'reg_id':     sanitize_excel_input(str(row[0])),
                    'name':       sanitize_excel_input(str(row[2]) if row[2] else ''),
                    'department': sanitize_excel_input(str(row[3]) if row[3] else ''),
                    'device_type': sanitize_excel_input(str(row[5]) if row[5] else ''),
                    'device':     sanitize_excel_input(str(row[6]) if row[6] else ''),
                    'serial':     sanitize_excel_input(str(row[9]) if row[9] else ''),
                    'status':     sanitize_excel_input(str(row[14]) if row[14] else 'Pending'),
                }
                wb.close()
                return device
        
        wb.close()
        return None
        
    except Exception as e:
        app.logger.error(f"[ERROR] Reading Excel: {e}")
        return None


def check_it_compliance(reg_id: str) -> bool:
    """Check if device has passed IT inspection."""
    try:
        wb = safe_load_workbook()
        sheet = wb['IT Inspection']
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and str(row[0]).strip() == reg_id.strip():
                compliance = str(row[13]).strip().lower() if row[13] else ''
                wb.close()
                return compliance == 'compliant'
        
        wb.close()
        return False
        
    except Exception as e:
        app.logger.error(f"[ERROR] Checking compliance: {e}")
        return False


def log_gate_entry(reg_id: str, name: str, device: str, action: str,
                   officer: str, status: str, remarks: str = '') -> bool:
    """
    SECURITY: P0 - Log with sanitized inputs to prevent formula injection.
    Append to Security Gate Log sheet.
    """
    try:
        # Sanitize all inputs
        reg_id = sanitize_excel_input(reg_id, 50)
        name = sanitize_excel_input(name, 100)
        device = sanitize_excel_input(device, 100)
        action = sanitize_excel_input(action, 20)
        officer = sanitize_excel_input(officer, 100)
        remarks = sanitize_excel_input(remarks, 255)
        
        wb = safe_load_workbook()
        sheet = wb['Security Gate Log']
        
        next_row = sheet.max_row + 1
        now = datetime.now()
        log_id = f"LOG-{now.strftime('%Y%m%d')}-{next_row - 1:04d}"
        
        row_data = [
            log_id,                      # A: Log ID
            now.strftime('%Y-%m-%d'),    # B: Date
            now.strftime('%H:%M:%S'),    # C: Time
            reg_id,                      # D: Registration ID
            name,                        # E: Name
            device,                      # F: Device
            action,                      # G: Action
            officer,                     # H: Officer
            status,                      # I: Status
            remarks                      # J: Remarks
        ]
        
        for col, value in enumerate(row_data, 1):
            sheet.cell(row=next_row, column=col).value = value
        
        if not safe_save_workbook(wb):
            app.logger.error(f"[ERROR] Failed to save log for {reg_id}")
            return False
        
        app.logger.info(f"[LOG] {log_id} | {action} | {name} | {status}")
        return True
        
    except Exception as e:
        app.logger.error(f"[ERROR] Writing gate log: {e}")
        return False


# ── ROUTES ──────────────────────────────────────────────────────────────────

@app.route('/gate/ping', methods=['GET'])
def ping():
    """Health check endpoint."""
    return jsonify({
        'ok': True,
        'server': 'NITDA Gate Server',
        'time': datetime.now().isoformat(),
        'auth_required': bool(GATE_API_KEY)
    })


@app.route('/gate/scan', methods=['POST'])
@require_api_key
@require_ip_whitelist
def gate_scan():
    """
    Main scan endpoint.
    
    Request (JSON):
        { "reg_id": "BYOD-...", "action": "in"|"out", "officer": "Security Officer" }
    
    Response:
        { "ok": true, "name": "...", "device": "...", ... }
        or
        { "ok": false, "reason": "...", "message": "..." }
    """
    
    # Rate limiting check
    client_ip = request.headers.get('X-Forwarded-For', request.remote_addr).split(',')[0].strip()
    if not check_rate_limit(client_ip):
        app.logger.warning(f"[SECURITY] Rate limit exceeded for {client_ip}")
        return jsonify({
            'ok': False,
            'reason': 'RATE_LIMITED',
            'message': 'Too many requests. Please wait before scanning again.'
        }), 429
    
    # Parse and validate request
    body = request.get_json(silent=True)
    if not body:
        return jsonify({
            'ok': False,
            'reason': 'BAD_REQUEST',
            'message': 'Invalid JSON request'
        }), 400
    
    reg_id = body.get('reg_id', '').strip()
    action = body.get('action', 'in').lower()
    officer = body.get('officer', 'Security Officer').strip()
    
    # Validate inputs
    if not validate_input(reg_id, r'^[A-Z0-9\-]{6,50}$', 50):
        log_gate_entry('INVALID', 'INVALID', '—', 'Scan', officer, 'Denied', 
                       'Invalid reg_id format')
        return jsonify({
            'ok': False,
            'reason': 'INVALID_FORMAT',
            'message': 'Invalid registration ID format'
        }), 400
    
    if action not in ('in', 'out'):
        return jsonify({
            'ok': False,
            'reason': 'INVALID_ACTION',
            'message': 'Action must be "in" or "out"'
        }), 400
    
    if not validate_input(officer, max_length=100):
        officer = 'Security Officer'
    
    action_label = 'Check In' if action == 'in' else 'Check Out'
    
    # Lookup device
    device = get_device_record(reg_id)
    
    if not device:
        log_gate_entry(reg_id, 'UNKNOWN', '—', action_label, officer,
                       'Denied', 'Registration ID not found')
        return jsonify({
            'ok': False,
            'reason': 'NOT_FOUND',
            'message': f'No device with ID: {reg_id}. Contact BYOD administrator.'
        })
    
    # Check approval status
    if device['status'].lower() != 'approved':
        log_gate_entry(reg_id, device['name'], device['device'], action_label,
                       officer, 'Denied', f"Status: {device['status']}")
        return jsonify({
            'ok': False,
            'reason': f"DEVICE_{device['status'].upper()}",
            'message': f"Status: {device['status']}. Must be Approved to proceed."
        })
    
    # Check IT compliance
    if not check_it_compliance(reg_id):
        log_gate_entry(reg_id, device['name'], device['device'], action_label,
                       officer, 'Denied', 'IT inspection not completed')
        return jsonify({
            'ok': False,
            'reason': 'INSPECTION_INCOMPLETE',
            'message': 'Device has not completed IT security inspection.'
        })
    
    # All checks passed
    now = datetime.now()
    logged = log_gate_entry(
        reg_id, device['name'], device['device'],
        action_label, officer, 'Authorised'
    )
    
    return jsonify({
        'ok': True,
        'reg_id': device['reg_id'],
        'name': device['name'],
        'department': device['department'],
        'device': device['device'],
        'serial': device['serial'],
        'status': device['status'],
        'action': action_label,
        'time': now.strftime('%H:%M:%S'),
        'date': now.strftime('%Y-%m-%d'),
        'logged': logged
    })


@app.route('/gate/log', methods=['GET'])
@require_api_key
@require_ip_whitelist
def get_todays_log():
    """Get today's gate log entries."""
    target_date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    
    # Validate date format
    try:
        datetime.strptime(target_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({'ok': False, 'message': 'Invalid date format'}), 400
    
    try:
        wb = safe_load_workbook()
        sheet = wb['Security Gate Log']
        entries = []
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and str(row[1]) == target_date:
                entries.append({
                    'log_id': row[0],
                    'time': str(row[2]),
                    'name': row[4],
                    'device': row[5],
                    'action': row[6],
                    'status': row[8]
                })
        
        wb.close()
        return jsonify({
            'ok': True,
            'date': target_date,
            'entries': entries,
            'count': len(entries)
        })
        
    except Exception as e:
        app.logger.error(f"[ERROR] Reading log: {e}")
        return jsonify({'ok': False, 'message': 'Error reading log'}), 500


# ── STARTUP ─────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    print("\n" + "=" * 70)
    print("  NITDA BYOD — SECURITY GATE SERVER")
    print("=" * 70)
    
    if not os.path.exists(EXCEL_FILE):
        print(f"\n  [ERROR] '{EXCEL_FILE}' not found!")
        exit(1)
    
    print(f"\n  Database     : {os.path.abspath(EXCEL_FILE)}")
    print(f"  API Key      : {'✓ CONFIGURED' if GATE_API_KEY else '✗ NOT SET (set GATE_API_KEY in .env)'}")
    print(f"  Allowed IPs  : {', '.join(ALLOWED_IPS)}")
    print(f"  Rate Limit   : {RATE_LIMIT_REQUESTS} req/{RATE_LIMIT_WINDOW}s per IP")
    print(f"\n  UI: http://localhost:5001/gate_scanner_prototype.html")
    print("=" * 70 + "\n")
    
    app.run(host='0.0.0.0', port=5001, debug=False)
