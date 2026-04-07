import openpyxl

def check_approved_devices():
    excel_file = 'NITDA_BYOD_Database.xlsx'
    
    print("\n" + "="*70)
    print("APPROVED DEVICES DIAGNOSTIC")
    print("="*70)
    
    # Open workbook
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb['Device Registrations']
    
    print("\nChecking Device Registrations sheet...")
    print("-" * 70)
    
    approved_count = 0
    pending_count = 0
    other_count = 0
    
    for row in sheet.iter_rows(min_row=2, values_only=False):
        reg_id = row[0].value  # Column A
        name = row[2].value     # Column C
        status = row[14].value  # Column O (index 14)
        
        if not reg_id:
            continue
        
        print(f"\nRow {row[0].row}:")
        print(f"  Registration ID: '{reg_id}'")
        print(f"  Name: '{name}'")
        print(f"  Status (Column O): '{status}'")
        print(f"  Status Type: {type(status)}")
        print(f"  Status == 'Approved': {status == 'Approved'}")
        print(f"  Status == 'Pending': {status == 'Pending'}")
        
        # Check what the automation script sees
        if status == 'Approved':
            approved_count += 1
            print(f"  ✓ Would be sent to IT inspection")
        elif status == 'Pending' or not status:
            pending_count += 1
            print(f"  → Would be processed as new registration")
        else:
            other_count += 1
            print(f"  ✗ Would be skipped (status: {status})")
    
    print("\n" + "="*70)
    print("SUMMARY:")
    print(f"  Approved devices (should go to IT): {approved_count}")
    print(f"  Pending devices (new registrations): {pending_count}")
    print(f"  Other status: {other_count}")
    print("="*70)
    
    # Check IT Inspection sheet
    print("\nChecking IT Inspection sheet...")
    print("-" * 70)
    
    if 'IT Inspection' in wb.sheetnames:
        it_sheet = wb['IT Inspection']
        print(f"✓ IT Inspection sheet exists")
        print(f"  Current rows: {it_sheet.max_row - 1} (excluding header)")
        
        print("\nExisting inspection records:")
        for row in it_sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # If Registration ID exists
                print(f"  - {row[0]}: {row[1]} (Status: {row[13]})")
    else:
        print("✗ IT Inspection sheet not found!")
    
    print("\n" + "="*70)
    
    if approved_count == 0:
        print("\n⚠️ PROBLEM: No approved devices found!")
        print("\nTo fix:")
        print("1. Open NITDA_BYOD_Database.xlsx")
        print("2. Go to 'Device Registrations' sheet")
        print("3. In Column O (Status), type exactly: Approved")
        print("   (Capital A, no extra spaces)")
        print("4. Fill in Column P (Approved By) and Q (Approval Date)")
        print("5. Save the file and close Excel completely")
        print("6. Run: python byod_automation.py")
    else:
        print(f"\n✓ Found {approved_count} approved device(s)")
        print("These should trigger IT inspection scheduling.")
        print("\nIf automation still doesn't schedule inspections, check:")
        print("1. Make sure Excel file is saved and closed")
        print("2. Check if IT inspection already exists for these devices")

if __name__ == "__main__":
    check_approved_devices()