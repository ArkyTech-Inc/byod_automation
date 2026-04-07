# ==================== This is the main script for creating the BYOD sheet on excel ====================

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# This Create/initializes the workbook
wb = Workbook()

# This remove default sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Adds Professional styling
header_fill = PatternFill(start_color='00A86B', end_color='00A86B', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ==================== SHEET 1: DEVICE REGISTRATIONS ====================
sheet1 = wb.create_sheet('Device Registrations', 0)

# Headers
headers1 = [
    'Registration ID', 'Timestamp', 'Name', 'Department', 'Classification',
    'Device Type', 'Device Make/Model', 'Operating System', 'MAC Address',
    'Serial Number', 'Email', 'Phone', 'Supervisor Name', 'Supervisor Email',
    'Status', 'Approved By', 'Approval Date', 'Admin Remarks'
]

for col_num, header in enumerate(headers1, 1):
    cell = sheet1.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border


# Column widths
column_widths1 = [15, 18, 20, 12, 12, 12, 18, 15, 18, 15, 25, 15, 20, 25, 12, 15, 15, 20]
for i, width in enumerate(column_widths1, 1):
    sheet1.column_dimensions[get_column_letter(i)].width = width

# ==================== SHEET 2: IT INSPECTION ====================
sheet2 = wb.create_sheet('IT Inspection', 1)

headers2 = [
    'Registration ID', 'Name', 'Device Model', 'Serial Number', 'Inspection ID',
    'Inspection Date', 'Inspected By', 'OS Updated', 'Anti-malware Installed',
    'Device Encrypted', 'Remote Wipe Enabled', 'Secure Access Configured',
    'No Unauthorized Apps', 'Compliance Status', 'Remarks', 'QR Code Generated',
    'Pass Issued Date'
]

for col_num, header in enumerate(headers2, 1):
    cell = sheet2.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Column widths
column_widths2 = [15, 20, 18, 15, 15, 15, 20, 12, 12, 12, 12, 12, 12, 15, 25, 15, 15]
for i, width in enumerate(column_widths2, 1):
    sheet2.column_dimensions[get_column_letter(i)].width = width

# ==================== SHEET 3: SECURITY GATE LOG ====================
sheet3 = wb.create_sheet('Security Gate Log', 2)

headers3 = [
    'Log ID', 'Date', 'Time', 'Registration ID', 'Name', 'Device Model',
    'Action', 'Security Officer', 'Status', 'Remarks'
]

for col_num, header in enumerate(headers3, 1):
    cell = sheet3.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Column widths
column_widths3 = [12, 12, 12, 15, 20, 18, 12, 20, 12, 25]
for i, width in enumerate(column_widths3, 1):
    sheet3.column_dimensions[get_column_letter(i)].width = width

# ==================== SHEET 4: DASHBOARD ====================
sheet4 = wb.create_sheet('Dashboard', 3)

# Title
sheet4['A1'] = 'NITDA BYOD MANAGEMENT DASHBOARD'
sheet4['A1'].font = Font(bold=True, size=16, color='00A86B')
sheet4.merge_cells('A1:F1')
sheet4['A1'].alignment = Alignment(horizontal='center', vertical='center')

# Key Metrics
metrics_start_row = 3
metrics = [
    ('Total Registrations', '=COUNTA(\'Device Registrations\'!A2:A1000)'),
    ('Pending Approval', '=COUNTIF(\'Device Registrations\'!O2:O1000,"Pending")'),
    ('Approved Devices', '=COUNTIF(\'Device Registrations\'!O2:O1000,"Approved")'),
    ('Rejected Devices', '=COUNTIF(\'Device Registrations\'!O2:O1000,"Rejected")'),
    ('IT Inspections Completed', '=COUNTA(\'IT Inspection\'!A2:A1000)'),
    ('Compliant Devices', '=COUNTIF(\'IT Inspection\'!N2:N1000,"Compliant")'),
    ('Non-Compliant Devices', '=COUNTIF(\'IT Inspection\'!N2:N1000,"Non-Compliant")'),
    ('QR Codes Generated', '=COUNTIF(\'IT Inspection\'!P2:P1000,"Yes")'),
    ('Total Gate Logs Today', '=COUNTIFS(\'Security Gate Log\'!B2:B1000,TODAY())'),
    ('Devices Currently Inside', '=COUNTIFS(\'Security Gate Log\'!B2:B1000,TODAY(),\'Security Gate Log\'!G2:G1000,"Check In")-COUNTIFS(\'Security Gate Log\'!B2:B1000,TODAY(),\'Security Gate Log\'!G2:G1000,"Check Out")')
]

for i, (label, formula) in enumerate(metrics, metrics_start_row):
    # Label
    cell_label = sheet4.cell(row=i, column=1)
    cell_label.value = label
    cell_label.font = Font(bold=True, size=11)
    cell_label.alignment = Alignment(horizontal='left', vertical='center')
    
    # Value
    cell_value = sheet4.cell(row=i, column=2)
    cell_value.value = formula
    cell_value.font = Font(bold=True, size=14, color='00A86B')
    cell_value.alignment = Alignment(horizontal='center', vertical='center')
    cell_value.fill = PatternFill(start_color='E8F5F0', end_color='E8F5F0', fill_type='solid')
    cell_value.border = thin_border

# Status breakdown
sheet4['A15'] = 'DEPARTMENT BREAKDOWN'
sheet4['A15'].font = Font(bold=True, size=12, color='00A86B')

departments = ['DLCB', 'ITIS', 'Admin', 'Finance', 'HR', 'Legal', 'EGDED']
sheet4['A16'] = 'Department'
sheet4['B16'] = 'Total Devices'
sheet4['A16'].font = Font(bold=True)
sheet4['B16'].font = Font(bold=True)

for i, dept in enumerate(departments, 17):
    sheet4[f'A{i}'] = dept
    sheet4[f'B{i}'] = f'=COUNTIF(\'Device Registrations\'!D2:D1000,"{dept}")'
    sheet4[f'B{i}'].alignment = Alignment(horizontal='center')

# Column widths
sheet4.column_dimensions['A'].width = 30
sheet4.column_dimensions['B'].width = 20

# ==================== SHEET 5: AUTOMATION SETTINGS ====================
sheet5 = wb.create_sheet('Automation Settings', 4)

sheet5['A1'] = 'AUTOMATION CONFIGURATION'
sheet5['A1'].font = Font(bold=True, size=14, color='00A86B')
sheet5.merge_cells('A1:C1')

settings = [
    ('Email Settings', '', ''),
    ('SMTP Server', 'smtp.gmail.com', 'Gmail SMTP server'),
    ('SMTP Port', '465', 'TLS port'),
    ('Sender Email', 'motolanisomoye@gmail.com', 'System email address'),
    ('', '', ''),
    ('IT Department Settings', '', ''),
    ('IT Email', 'somoye2000@gmail.com', 'IT inspection notifications'),
    ('Inspection Lead Time (days)', '2', 'Days to schedule inspection'),
    ('', '', ''),
    ('QR Code Settings', '', ''),
    ('QR Code Size', '300', 'Pixels'),
    ('QR Code Error Correction', 'H', 'High error correction'),
    ('', '', ''),
    ('Auto-Approval Settings', '', ''),
    ('Auto-approve Staff', 'No', 'Auto-approve staff devices'),
    ('Require Supervisor Endorsement', 'Yes', 'Supervisor must endorse registrations'),
]

row = 3
for setting, value, description in settings:
    sheet5[f'A{row}'] = setting
    sheet5[f'B{row}'] = value
    sheet5[f'C{row}'] = description
    
    if setting and not setting.endswith('Settings'):
        sheet5[f'A{row}'].font = Font(bold=False)
        sheet5[f'B{row}'].fill = PatternFill(start_color='E8F5F0', end_color='E8F5F0', fill_type='solid')
    else:
        sheet5[f'A{row}'].font = Font(bold=True, size=11)
    
    row += 1

sheet5.column_dimensions['A'].width = 30
sheet5.column_dimensions['B'].width = 25
sheet5.column_dimensions['C'].width = 35

# ==================== SHEET 6: EMAIL TEMPLATES ====================
sheet6 = wb.create_sheet('Email Templates', 5)

sheet6['A1'] = 'EMAIL TEMPLATES'
sheet6['A1'].font = Font(bold=True, size=14, color='00A86B')

templates = [
    ('Template Name', 'Subject', 'Body'),
    ('Registration Confirmation', 
     'BYOD Registration Received - {registration_id}',
     '''Dear {name},

Your BYOD registration has been received successfully.

Registration ID: {registration_id}
Device: {device_model}
Date: {date}

Your application is now pending supervisor endorsement. You will receive a notification once your supervisor reviews your application.

Please ensure your device meets all security requirements as outlined in the BYOD policy.

Best regards,
NITDA Admin Team'''),
    
    ('Supervisor Approval Request',
     'BYOD Endorsement Required - {name}',
     '''Dear {supervisor_name},

A new BYOD registration requires your endorsement.

Intern's Name: {name}
Department: {department}
Device: {device_model}
Registration ID: {registration_id}

Please review and endorse/reject this registration in the BYOD system.

Link: [Dashboard Link]

Best regards,
NITDA BYOD System'''),
    
    ('Admin Approval Notification',
     'BYOD Approved - Schedule IT Inspection',
     '''Dear {name},

Your BYOD registration has been endorsed by your supervisor.

Registration ID: {registration_id}
Device: {device_model}

Next Steps:
1. The IT department will contact you within 2 business days
2. Bring your device for security compliance inspection
3. Upon passing inspection, you will receive your QR code pass

Best regards,
NITDA Admin Team'''),
    
    ('IT Inspection Schedule',
     'BYOD Security Inspection - {registration_id}',
     '''Dear {name},

Please bring your device for security compliance inspection.

Registration ID: {registration_id}
Device: {device_model}
Scheduled Date: {inspection_date}
Location: IT Department, 2nd Floor

Requirements to Check:
- OS is updated
- Anti-malware installed
- Device encryption enabled
- Screen lock configured
- No unauthorized apps

Please contact IT at motolanisomoye@gmail.com if you need to reschedule.

Best regards,
NITDA IT Department'''),
    
    ('QR Code Pass Issued',
     'BYOD Pass Issued - {registration_id}',
     '''Dear {name},

Congratulations! Your device has passed security inspection.

Registration ID: {registration_id}
Device: {device_model}
Inspection Status: COMPLIANT

Your QR code pass is attached to this email. Please:
1. Save the QR code on your phone
2. Print a copy for backup
3. Present the QR code at security gate when bringing your device into the office.
   The QR code contains all your device information for quick verification,
   you can sign in via qr code or your registation code at the security gate.

The QR code contains all your device information for quick verification, keep it safe.

Best regards,
NITDA IT Department'''),
]

row = 3
for template_data in templates:
    for col, value in enumerate(template_data, 1):
        cell = sheet6.cell(row=row, column=col)
        cell.value = value
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        if row == 3:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.font = header_font
        cell.border = thin_border
    row += 1

sheet6.column_dimensions['A'].width = 25
sheet6.column_dimensions['B'].width = 40
sheet6.column_dimensions['C'].width = 60
sheet6.row_dimensions[4].height = 150
sheet6.row_dimensions[5].height = 100
sheet6.row_dimensions[6].height = 120
sheet6.row_dimensions[7].height = 150
sheet6.row_dimensions[8].height = 150

# Saves workbook to file
output_file = 'NITDA_BYOD_Database.xlsx'
wb.save(output_file)
print(f"Excel database created successfully: {output_file}")
print(f"Saved in: {os.path.abspath(output_file)}")
