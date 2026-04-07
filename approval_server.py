"""
BYOD Approval Web Server
This creates a simple web server that handles approve/reject links from emails
"""

from flask import Flask, request, render_template_string
import openpyxl
from datetime import datetime
import os

app = Flask(__name__)

# HTML templates
APPROVAL_PAGE = """
<!DOCTYPE html>
<html>
<head>
    <title>NITDA BYOD Approval</title>
    <style>
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #f5f7fa 0%, #e8f5f0 100%);
            margin: 0;
            padding: 20px;
            display: flex;
            justify-content: center;
            align-items: center;
            min-height: 100vh;
        }
        .container {
            background: white;
            padding: 40px;
            border-radius: 16px;
            box-shadow: 0 10px 40px rgba(0, 0, 0, 0.1);
            max-width: 600px;
            width: 100%;
        }
        h1 {
            color: #00A86B;
            margin-bottom: 10px;
        }
        .subtitle {
            color: #666;
            margin-bottom: 30px;
        }
        .device-info {
            background: #f8f9fa;
            padding: 20px;
            border-radius: 8px;
            margin: 20px 0;
            border-left: 4px solid #00A86B;
        }
        .info-row {
            display: flex;
            padding: 8px 0;
            border-bottom: 1px solid #e0e0e0;
        }
        .info-row:last-child {
            border-bottom: none;
        }
        .label {
            font-weight: 600;
            color: #666;
            width: 150px;
        }
        .value {
            color: #1a1a1a;
            flex: 1;
        }
        .buttons {
            display: flex;
            gap: 15px;
            margin-top: 30px;
        }
        button {
            flex: 1;
            padding: 15px 30px;
            border: none;
            border-radius: 8px;
            font-size: 16px;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.3s ease;
        }
        .approve-btn {
            background: #00A86B;
            color: white;
        }
        .approve-btn:hover {
            background: #007850;
            transform: translateY(-2px);
            box-shadow: 0 4px 15px rgba(0, 168, 107, 0.3);
        }
        .reject-btn {
            background: #FF6B35;
            color: white;
        }
        .reject-btn:hover {
            background: #E85A2A;
            transform: translateY(-2px);
            box-shadow: 0 4px 15px rgba(255, 107, 53, 0.3);
        }
        .remarks {
            margin-top: 20px;
        }
        textarea {
            width: 100%;
            padding: 12px;
            border: 2px solid #e0e0e0;
            border-radius: 8px;
            font-family: inherit;
            font-size: 14px;
            resize: vertical;
            min-height: 80px;
        }
        textarea:focus {
            outline: none;
            border-color: #00A86B;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>Device Registration Approval</h1>
        <p class="subtitle">National Information Technology Development Agency</p>
        
        <div class="device-info">
            <div class="info-row">
                <span class="label">Registration ID:</span>
                <span class="value">{{ reg_id }}</span>
            </div>
            <div class="info-row">
                <span class="label">Name:</span>
                <span class="value">{{ name }}</span>
            </div>
            <div class="info-row">
                <span class="label">Department:</span>
                <span class="value">{{ department }}</span>
            </div>
            <div class="info-row">
                <span class="label">Device:</span>
                <span class="value">{{ device }}</span>
            </div>
            <div class="info-row">
                <span class="label">Serial Number:</span>
                <span class="value">{{ serial }}</span>
            </div>
        </div>
        
        <form method="POST">
            <div class="remarks">
                <label for="remarks" style="font-weight: 600; color: #666; display: block; margin-bottom: 8px;">
                    Remarks (Optional):
                </label>
                <textarea name="remarks" id="remarks" placeholder="Add any comments..."></textarea>
            </div>
            
            <div class="buttons">
                <button type="submit" name="action" value="approve" class="approve-btn">
                    ✓ Approve Device
                </button>
                <button type="submit" name="action" value="reject" class="reject-btn">
                    ✗ Reject Device
                </button>
            </div>
        </form>
    </div>
</body>
</html>
"""

SUCCESS_PAGE = """
<!DOCTYPE html>
<html>
<head>
    <title>Action Completed</title>
    <style>
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #f5f7fa 0%, #e8f5f0 100%);
            margin: 0;
            padding: 20px;
            display: flex;
            justify-content: center;
            align-items: center;
            min-height: 100vh;
        }
        .container {
            background: white;
            padding: 60px 40px;
            border-radius: 16px;
            box-shadow: 0 10px 40px rgba(0, 0, 0, 0.1);
            max-width: 500px;
            text-align: center;
        }
        .icon {
            font-size: 64px;
            margin-bottom: 20px;
        }
        .success { color: #00A86B; }
        .error { color: #FF6B35; }
        h1 {
            color: #1a1a1a;
            margin-bottom: 10px;
        }
        p {
            color: #666;
            font-size: 16px;
            line-height: 1.6;
        }
        .details {
            background: #f8f9fa;
            padding: 20px;
            border-radius: 8px;
            margin-top: 20px;
            text-align: left;
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="icon {{ status_class }}">{{ icon }}</div>
        <h1>{{ title }}</h1>
        <p>{{ message }}</p>
        {% if details %}
        <div class="details">
            <strong>Registration ID:</strong> {{ details.reg_id }}<br>
            <strong>Name:</strong> {{ details.name }}<br>
            <strong>Action:</strong> {{ details.action }}<br>
            <strong>By:</strong> {{ details.by }}<br>
            <strong>Date:</strong> {{ details.date }}
        </div>
        {% endif %}
    </div>
</body>
</html>
"""

def get_device_info(reg_id):
    """Get device information from Excel"""
    try:
        wb = openpyxl.load_workbook('NITDA_BYOD_Database.xlsx')
        sheet = wb['Device Registrations']
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == reg_id:
                return {
                    'reg_id': row[0],
                    'name': row[2],
                    'department': row[3],
                    'device': row[6],
                    'serial': row[9],
                    'email': row[10]
                }
        return None
    except Exception as e:
        print(f"Error reading device info: {e}")
        return None

def update_device_status(reg_id, action, remarks, approver):
    """Update device status in Excel"""
    try:
        wb = openpyxl.load_workbook('NITDA_BYOD_Database.xlsx')
        sheet = wb['Device Registrations']
        
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == reg_id:
                # Update status
                row[14].value = 'Approved' if action == 'approve' else 'Rejected'  # Column O
                row[15].value = approver  # Column P - Approved By
                row[16].value = datetime.now().strftime('%Y-%m-%d')  # Column Q - Approval Date
                row[17].value = remarks if remarks else ''  # Column R - Admin Remarks
                
                wb.save('NITDA_BYOD_Database.xlsx')
                return True
        
        return False
    except Exception as e:
        print(f"Error updating status: {e}")
        return False

@app.route('/approve/<reg_id>')
def approval_page(reg_id):
    """Show approval page"""
    device = get_device_info(reg_id)
    
    if not device:
        return render_template_string(
            SUCCESS_PAGE,
            status_class='error',
            icon='⚠️',
            title='Device Not Found',
            message=f'No device found with Registration ID: {reg_id}',
            details=None
        )
    
    return render_template_string(
        APPROVAL_PAGE,
        reg_id=device['reg_id'],
        name=device['name'],
        department=device['department'],
        device=device['device'],
        serial=device['serial']
    )

@app.route('/approve/<reg_id>', methods=['POST'])
def process_approval(reg_id):
    """Process approval/rejection"""
    action = request.form.get('action')
    remarks = request.form.get('remarks', '')
    
    # Get device info
    device = get_device_info(reg_id)
    
    if not device:
        return render_template_string(
            SUCCESS_PAGE,
            status_class='error',
            icon='⚠️',
            title='Error',
            message='Device not found',
            details=None
        )
    
    # Update status
    approver = 'Supervisor'  # You can customize this
    success = update_device_status(reg_id, action, remarks, approver)
    
    if success:
        # Run automation to process the approval
        import subprocess
        subprocess.Popen(['python', 'byod_automation.py'])
        
        if action == 'approve':
            return render_template_string(
                SUCCESS_PAGE,
                status_class='success',
                icon='✓',
                title='Device Approved!',
                message='The device has been approved and the user will be notified. IT inspection will be scheduled automatically.',
                details={
                    'reg_id': reg_id,
                    'name': device['name'],
                    'action': 'Approved',
                    'by': approver,
                    'date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
            )
        else:
            return render_template_string(
                SUCCESS_PAGE,
                status_class='error',
                icon='✗',
                title='Device Rejected',
                message='The device registration has been rejected. The user will be notified.',
                details={
                    'reg_id': reg_id,
                    'name': device['name'],
                    'action': 'Rejected',
                    'by': approver,
                    'date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
            )
    else:
        return render_template_string(
            SUCCESS_PAGE,
            status_class='error',
            icon='⚠️',
            title='Error',
            message='Failed to update device status. Please try again.',
            details=None
        )

@app.route('/')
def index():
    """Home page"""
    return """
    <html>
    <head>
        <title>NITDA BYOD Approval System</title>
        <style>
            body {
                font-family: Arial, sans-serif;
                max-width: 600px;
                margin: 50px auto;
                padding: 20px;
                background: #f5f5f5;
            }
            .container {
                background: white;
                padding: 40px;
                border-radius: 8px;
                box-shadow: 0 2px 10px rgba(0,0,0,0.1);
            }
            h1 { color: #00A86B; }
            p { color: #666; line-height: 1.6; }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>NITDA BYOD Approval System</h1>
            <p>This is the approval server for the BYOD device registration system.</p>
            <p>Supervisors will receive approval links via email.</p>
            <p><strong>Status:</strong> Server is running ✓</p>
        </div>
    </body>
    </html>
    """

if __name__ == '__main__':
    print("\n" + "="*70)
    print("NITDA BYOD APPROVAL SERVER")
    print("="*70)
    print("\nServer starting on http://localhost:5000")
    print("Supervisors can click approval links in their emails")
    print("\nPress Ctrl+C to stop the server")
    print("="*70 + "\n")
    
    app.run(host='0.0.0.0', port=5000, debug=False)
